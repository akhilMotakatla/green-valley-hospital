"""
Green Valley Hospital — Scrum Tracker Excel Generator
Run this script to regenerate docs/scrum-tracker.xlsx from the latest project state.
Usage: python docs/generate_tracker.py
"""

import json
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, GradientFill, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, GradientFill, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter

# ── Colour palette ──────────────────────────────────────────────────────────
C_GREEN_DARK  = "0E5E52"   # header backgrounds
C_GREEN_MID   = "1A8A7A"   # sub-headers
C_GREEN_LIGHT = "E6F5F2"   # alternating rows / section bg
C_ACCENT      = "E05C2A"   # pending / blocked status
C_WHITE       = "FFFFFF"
C_GREY_LIGHT  = "F4F7F6"
C_GREY_MID    = "D8E3E0"
C_TEXT_DARK   = "1A2422"
C_DONE        = "1E8E5A"   # completed status
C_PROGRESS    = "1F5AA8"   # in-progress status
C_PENDING     = "B7791F"   # pending status

STATUS_COLORS = {
    "Completed":   C_DONE,
    "In Progress": C_PROGRESS,
    "Pending":     C_PENDING,
    "Blocked":     C_ACCENT,
}


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=C_TEXT_DARK, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _border():
    thin = Side(style="thin", color=C_GREY_MID)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _header_row(ws, row, cols, bg=C_GREEN_DARK, fg=C_WHITE, size=11):
    for col, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _fill(bg)
        c.font = _font(bold=True, color=fg, size=size)
        c.alignment = _center(wrap=True)
        c.border = _border()


def _data_row(ws, row, cols, bg=C_WHITE, bold=False):
    for col, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _fill(bg)
        c.font = _font(bold=bold, color=C_TEXT_DARK)
        c.alignment = _left(wrap=True)
        c.border = _border()


def _status_cell(ws, row, col, status):
    c = ws.cell(row=row, column=col, value=status)
    color = STATUS_COLORS.get(status, C_TEXT_DARK)
    c.fill = _fill(C_WHITE)
    c.font = _font(bold=True, color=color)
    c.alignment = _center()
    c.border = _border()


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title_block(ws, title, subtitle, merge_cols=8):
    ws.merge_cells(f"A1:{get_column_letter(merge_cols)}1")
    t = ws["A1"]
    t.value = title
    t.fill = _fill(C_GREEN_DARK)
    t.font = Font(bold=True, color=C_WHITE, size=16, name="Calibri")
    t.alignment = _center()
    ws.row_dimensions[1].height = 40

    ws.merge_cells(f"A2:{get_column_letter(merge_cols)}2")
    s = ws["A2"]
    s.value = subtitle
    s.fill = _fill(C_GREEN_MID)
    s.font = Font(color=C_WHITE, size=11, name="Calibri")
    s.alignment = _center()
    ws.row_dimensions[2].height = 22


# ── Sheet 1: Project Overview ────────────────────────────────────────────────
def sheet_overview(wb):
    ws = wb.active
    ws.title = "Project Overview"
    ws.sheet_view.showGridLines = False

    _title_block(ws, "🏥  Green Valley Hospital — Project Overview",
                 f"Generated: {datetime.now().strftime('%B %d, %Y  %I:%M %p CST')}",
                 merge_cols=6)

    # Project details block
    details = [
        ("Project Name",   "Green Valley Hospital Management System"),
        ("Client",         "Green Valley Hospital"),
        ("Project Type",   "Full-Stack Web Application"),
        ("Tech Stack",     "Python (FastAPI) · React 19 (Vite + TypeScript) · SQLite · JWT Auth"),
        ("Methodology",    "Agile Scrum — Daily Standups"),
        ("Sprint",         "Sprint 1 — Visual Enhancement & Full-Stack Implementation"),
        ("Sprint Start",   "July 18, 2026"),
        ("Sprint Status",  "In Progress"),
        ("Scrum Master",   "Sunny"),
        ("Orchestrator",   "Akhil"),
    ]

    row = 4
    for label, value in details:
        ws.cell(row=row, column=1, value=label).fill = _fill(C_GREEN_LIGHT)
        ws.cell(row=row, column=1).font = _font(bold=True)
        ws.cell(row=row, column=1).border = _border()
        ws.cell(row=row, column=1).alignment = _left()

        ws.merge_cells(f"B{row}:F{row}")
        ws.cell(row=row, column=2, value=value).font = _font()
        ws.cell(row=row, column=2).border = _border()
        ws.cell(row=row, column=2).alignment = _left()
        ws.row_dimensions[row].height = 20
        row += 1

    # Overall pipeline status
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="PIPELINE STATUS").fill = _fill(C_GREEN_DARK)
    ws.cell(row=row, column=1).font = _font(bold=True, color=C_WHITE, size=12)
    ws.cell(row=row, column=1).alignment = _center()
    ws.row_dimensions[row].height = 28
    row += 1

    _header_row(ws, row, ["#", "Stage", "Agent", "Status", "Completion %", "Notes"], bg=C_GREEN_MID)
    ws.row_dimensions[row].height = 22
    row += 1

    pipeline = [
        (1, "Requirements Gathering",      "Lavanya",  "Completed",   "100%", "Section 6 Visual & UI Enhancement requirements written to docs/requirements.md"),
        (2, "System Architecture & Design","Sagar",    "Completed",   "100%", "docs/api-spec.md, docs/design.md, db/schema.sql updated with v1.1 changes"),
        (3, "Backend Implementation",      "Pavan",    "Completed",   "100%", "FastAPI endpoints updated: profile_photo_path, recent_articles, department response shape, seed data"),
        (4, "Frontend Implementation",     "Chintu",   "Completed",   "100%", "1891 modules, 0 TypeScript errors. All Section 6 visual pages, components, CSS tokens implemented"),
        (5, "QA & Testing",                "Gopal",    "In Progress", "60%",  "Running automated tests against Section 6 acceptance criteria + core API tests"),
        (6, "DevOps & Deployment",         "Indra",    "Pending",     "0%",   "Awaiting client go-ahead after QA passes"),
        (7, "Application Launch",          "—",        "Pending",     "0%",   "Final sign-off required from client"),
    ]

    for i, (num, stage, agent, status, pct, notes) in enumerate(pipeline):
        bg = C_GREEN_LIGHT if i % 2 == 0 else C_WHITE
        _data_row(ws, row, [num, stage, agent, "", pct, notes], bg=bg)
        _status_cell(ws, row, 4, status)
        ws.row_dimensions[row].height = 36
        row += 1

    _set_col_widths(ws, [5, 30, 14, 15, 14, 55])


# ── Sheet 2: Agent Status ────────────────────────────────────────────────────
def sheet_agents(wb):
    ws = wb.create_sheet("Agent Status")
    ws.sheet_view.showGridLines = False

    _title_block(ws, "👥  Agent Status & Deliverables",
                 f"Last Updated: {datetime.now().strftime('%B %d, %Y  %I:%M %p CST')}",
                 merge_cols=7)

    row = 4
    _header_row(ws, row,
                ["Agent", "Role", "Status", "Files Changed / Created", "Key Deliverables", "Bugs / Notes", "Completion"],
                bg=C_GREEN_DARK)
    ws.row_dimensions[row].height = 24
    row += 1

    agents = [
        (
            "Lavanya", "Requirements Analyst", "Completed",
            "docs/requirements.md",
            "• Added Section 6 (Visual & UI Enhancement) with 21 subsections\n"
            "• Defined 12 QA acceptance criteria (AC-VI-1 to AC-VI-12)\n"
            "• Specified 22 image assets, CSS token set, icon assignments\n"
            "• Backend change notes for Pavan: profile_photo_path, recent_articles, dept response",
            "None", "100%"
        ),
        (
            "Sagar", "Solution Architect", "Completed",
            "docs/api-spec.md\ndocs/design.md\ndb/schema.sql",
            "• Updated API spec: 5 endpoint changes + new Section 6 notes block\n"
            "• Added profile_photo_path to Doctor, cover_image_path to BlogArticle\n"
            "• Defined full-bleed layout strategy (page-root, no negative-margin)\n"
            "• Documented CSS token additions, image asset strategy, responsive breakpoints\n"
            "• Chose GET /public/home as source for recent_articles (no second API call)",
            "None", "100%"
        ),
        (
            "Pavan", "Backend Developer", "Completed",
            "src/backend/app/models.py\nsrc/backend/app/schemas.py\nsrc/backend/app/database.py\n"
            "src/backend/app/routers/public.py\nsrc/backend/app/routers/doctor.py\n"
            "src/backend/app/routers/patient.py\nsrc/backend/app/routers/staff.py\n"
            "db/seed/seed.py",
            "• profile_photo_path added to Doctor model (nullable, idempotent ALTER)\n"
            "• GET /public/home → recent_articles array + featured_departments.first_doctor\n"
            "• GET /public/departments/{id}/doctors → {department, items} shape\n"
            "• profile_photo_path included in all doctor endpoints (public, doctor, patient, staff)\n"
            "• PATCH /api/doctor/me accepts profile_photo_path\n"
            "• Seed: 4 doctors with photo paths, 3 extra blog articles, Pediatrics/Neurology depts",
            "None", "100%"
        ),
        (
            "Chintu", "Frontend Developer", "Completed",
            "src/frontend/src/index.css\nsrc/frontend/index.html\n"
            "src/frontend/src/components/* (4 new)\nsrc/frontend/src/layouts/*\n"
            "src/frontend/src/pages/public/* (9 pages)\nsrc/frontend/src/pages/admin/*\n"
            "src/frontend/src/pages/patient/*\nsrc/frontend/src/pages/doctor/*\n"
            "src/frontend/src/pages/staff/*\nsrc/frontend/src/pages/lab/*\n"
            "src/frontend/public/images/* (28 SVG placeholders)",
            "• Full CSS design token system (colors, shadows, radii, typography)\n"
            "• Inter font via Google Fonts, Lucide React icons installed\n"
            "• Logo, SkeletonBlock, PageError, BackToTopButton components\n"
            "• Emergency strip + 4-column footer + sticky nav with hamburger\n"
            "• Home: hero, stats counter animation, Why Choose Us, Specialists, Testimonials, Blog\n"
            "• About: timeline, facility gallery, mission/vision/values cards\n"
            "• Departments: hero + search filter + image cards\n"
            "• Doctor profile: two-column layout + booking CTA strip\n"
            "• Contact: split layout + map placeholder + success state\n"
            "• Auth: split-screen login/signup with icon inputs + password toggle\n"
            "• All dashboards: sidebar icons, user block, topbar, skeleton loaders\n"
            "• Build: 1891 modules, 0 TypeScript errors, 799ms",
            "None — clean build", "100%"
        ),
        (
            "Gopal", "QA Engineer", "In Progress",
            "tests/* (new test files)\ndocs/qa-report.md",
            "• Running existing test suite\n"
            "• Writing new tests: Section 6 backend endpoint shapes\n"
            "• Writing tests: auth, patient isolation, doctor-patient RBAC, lab workflow\n"
            "• Generating QA report to docs/qa-report.md",
            "TBD — in progress", "60%"
        ),
        (
            "Indra", "DevOps Engineer", "Pending",
            "—",
            "• Wire up run scripts for backend + frontend\n"
            "• Seed database and verify app boots end-to-end\n"
            "• Confirm all API endpoints reachable\n"
            "• Verify image assets served correctly",
            "Awaiting client go-ahead", "0%"
        ),
    ]

    for i, row_data in enumerate(agents):
        bg = C_GREEN_LIGHT if i % 2 == 0 else C_WHITE
        agent, role, status, files, deliverables, bugs, pct = row_data
        _data_row(ws, row, [agent, role, "", files, deliverables, bugs, pct], bg=bg)
        _status_cell(ws, row, 3, status)
        ws.row_dimensions[row].height = 100
        row += 1

    _set_col_widths(ws, [14, 22, 14, 35, 60, 28, 13])


# ── Sheet 3: Daily Scrum Log ─────────────────────────────────────────────────
def sheet_scrum_log(wb):
    ws = wb.create_sheet("Daily Scrum Log")
    ws.sheet_view.showGridLines = False

    _title_block(ws, "📋  Daily Scrum Log — Green Valley Hospital",
                 "Managed by Sunny (Scrum Master) · 9am standup & 5pm update every weekday",
                 merge_cols=7)

    row = 4
    _header_row(ws, row,
                ["Date", "Standup", "Agent", "Yesterday / Plan", "Today's Plan / Completed", "Blockers", "Status"],
                bg=C_GREEN_DARK)
    ws.row_dimensions[row].height = 24
    row += 1

    today = datetime.now().strftime("%b %d, %Y")

    log_entries = [
        (today, "Morning (9am)", "Lavanya",
         "Sprint kick-off",
         "Gather all visual & UI enhancement requirements for the website. Define image strategy, color tokens, icon system, page-by-page redesign specs, and backend change notes.",
         "None", "Completed"),
        (today, "Evening (5pm)", "Lavanya",
         "—",
         "✅ Completed Section 6 of docs/requirements.md — 21 subsections covering all UI enhancements, 12 QA acceptance criteria, 22 image asset specs, CSS token additions, icon mapping table, backend change notes for Pavan.",
         "None", "Completed"),
        (today, "Morning (9am)", "Sagar",
         "Requirements from Lavanya received",
         "Design system architecture. Update API spec for profile_photo_path and dept response shape. Produce docs/design.md with layout strategy and component hierarchy.",
         "None", "Completed"),
        (today, "Evening (5pm)", "Sagar",
         "—",
         "✅ Updated docs/api-spec.md (5 endpoint changes + notes block). ✅ Added profile_photo_path to db/schema.sql. ✅ Created docs/design.md (16 sections) with full-bleed layout approach, CSS token spec, responsive breakpoints, image strategy.",
         "None", "Completed"),
        (today, "Morning (9am)", "Pavan",
         "Design artifacts from Sagar received",
         "Implement all backend changes: profile_photo_path on Doctor model, recent_articles on home endpoint, department object in dept-doctors endpoint, update all doctor endpoints.",
         "None", "Completed"),
        (today, "Morning (9am)", "Chintu",
         "Design artifacts from Sagar received",
         "Implement all frontend visual enhancements per Section 6. CSS tokens, icons, all public pages, auth pages, all dashboards, shared components, image placeholders.",
         "None", "Completed"),
        (today, "Evening (5pm)", "Pavan",
         "—",
         "✅ 8 files updated. Doctor model gains profile_photo_path. All 6 doctor-related endpoints updated. Home endpoint returns recent_articles. Dept-doctors returns {department, items}. Seed data updated with 4 doctors + photo paths + 3 blog articles.",
         "None", "Completed"),
        (today, "Evening (5pm)", "Chintu",
         "—",
         "✅ Build clean: 1891 modules, 0 TypeScript errors, 799ms. All 20+ pages redesigned. 4 new components. 28 SVG image placeholders. Full CSS token system. Lucide icons throughout. Counter animations, hamburger nav, split-screen auth, skeleton loaders, empty states.",
         "None", "Completed"),
        (today, "Morning (9am)", "Gopal",
         "Pavan + Chintu implementation complete",
         "Run existing test suite. Write new tests for Section 6 backend changes and core acceptance criteria. Generate QA report.",
         "None", "In Progress"),
    ]

    for i, (date, standup, agent, yesterday, today_col, blocker, status) in enumerate(log_entries):
        bg = C_GREEN_LIGHT if i % 2 == 0 else C_WHITE
        _data_row(ws, row, [date, standup, agent, yesterday, today_col, blocker, ""], bg=bg)
        _status_cell(ws, row, 7, status)
        ws.row_dimensions[row].height = 70
        row += 1

    # Empty rows for Sunny to fill going forward
    for i in range(10):
        bg = C_GREEN_LIGHT if (len(log_entries) + i) % 2 == 0 else C_WHITE
        _data_row(ws, row, ["", "", "", "", "", "", ""], bg=bg)
        ws.row_dimensions[row].height = 40
        row += 1

    _set_col_widths(ws, [14, 16, 14, 45, 60, 28, 14])


# ── Sheet 4: Deliverables ────────────────────────────────────────────────────
def sheet_deliverables(wb):
    ws = wb.create_sheet("Deliverables")
    ws.sheet_view.showGridLines = False

    _title_block(ws, "📦  Project Deliverables",
                 "All files, documents, and artifacts produced in this sprint",
                 merge_cols=5)

    row = 4
    _header_row(ws, row, ["Category", "File / Artifact", "Owner", "Status", "Description"], bg=C_GREEN_DARK)
    ws.row_dimensions[row].height = 24
    row += 1

    deliverables = [
        # Documentation
        ("Documentation", "docs/requirements.md",    "Lavanya", "Completed", "Full functional requirements including Section 6 visual UI enhancements"),
        ("Documentation", "docs/api-spec.md",        "Sagar",   "Completed", "REST API specification with all endpoint schemas (v1.1)"),
        ("Documentation", "docs/design.md",          "Sagar",   "Completed", "System architecture, component hierarchy, CSS strategy, layout decisions"),
        ("Documentation", "docs/architecture.md",    "Sagar",   "Completed", "High-level system architecture overview"),
        ("Documentation", "docs/data-dictionary.md", "Sunny",   "Completed", "All 16 database entities with column descriptions"),
        ("Documentation", "docs/developer-guide.md", "Sunny",   "Completed", "Setup, installation, running, seeding, demo accounts, troubleshooting"),
        ("Documentation", "docs/user-guide.md",      "Sunny",   "Completed", "Role-by-role feature walkthroughs for all 5 roles + public site"),
        ("Documentation", "docs/security.md",        "Sunny",   "Completed", "JWT auth, RBAC, access control matrix, password policy, out-of-scope items"),
        ("Documentation", "docs/qa-report.md",       "Gopal",   "In Progress","QA test results and bug log"),
        ("Documentation", "docs/scrum-tracker.xlsx", "Sunny",   "Completed", "This file — daily scrum log, agent status, pipeline tracker"),
        # Database
        ("Database",      "db/schema.sql",           "Sagar",   "Completed", "Canonical DDL (v1.1): added profile_photo_path to doctors, cover_image_path to blog_articles"),
        ("Database",      "db/seed/seed.py",         "Pavan",   "Completed", "Idempotent seed: 4 doctors with photo paths, 8 departments, 3 blog articles, all roles"),
        # Backend
        ("Backend",       "src/backend/app/models.py",          "Pavan", "Completed", "SQLAlchemy models — Doctor gains profile_photo_path"),
        ("Backend",       "src/backend/app/schemas.py",         "Pavan", "Completed", "Pydantic schemas — DoctorUpdateMeRequest accepts profile_photo_path"),
        ("Backend",       "src/backend/app/database.py",        "Pavan", "Completed", "init_db with idempotent ALTER TABLE migration"),
        ("Backend",       "src/backend/app/routers/public.py",  "Pavan", "Completed", "Home returns recent_articles; dept-doctors returns {department, items}"),
        ("Backend",       "src/backend/app/routers/doctor.py",  "Pavan", "Completed", "GET/PATCH /doctor/me include profile_photo_path"),
        ("Backend",       "src/backend/app/routers/patient.py", "Pavan", "Completed", "GET /patients/doctors includes profile_photo_path"),
        ("Backend",       "src/backend/app/routers/staff.py",   "Pavan", "Completed", "GET /staff/directory includes profile_photo_path"),
        # Frontend
        ("Frontend",      "src/frontend/src/index.css",                    "Chintu", "Completed", "Full design token system, all component styles, responsive breakpoints"),
        ("Frontend",      "src/frontend/index.html",                       "Chintu", "Completed", "Inter font via Google Fonts preconnect"),
        ("Frontend",      "src/frontend/src/components/Logo.tsx",          "Chintu", "Completed", "SVG logo mark — default and white variants"),
        ("Frontend",      "src/frontend/src/components/SkeletonBlock.tsx", "Chintu", "Completed", "Shimmer loading skeleton — replaces all plain 'Loading...' text"),
        ("Frontend",      "src/frontend/src/components/PageError.tsx",     "Chintu", "Completed", "Error card with AlertCircle icon and retry button"),
        ("Frontend",      "src/frontend/src/components/BackToTopButton.tsx","Chintu","Completed", "Floating back-to-top button with ChevronUp icon"),
        ("Frontend",      "src/frontend/src/layouts/PublicLayout.tsx",     "Chintu", "Completed", "Emergency strip, Logo, hamburger menu, 4-column footer"),
        ("Frontend",      "src/frontend/src/layouts/AppShell.tsx",         "Chintu", "Completed", "Sidebar with icons, user block, Bell topbar"),
        ("Frontend",      "src/frontend/src/pages/public/HomePage.tsx",    "Chintu", "Completed", "Hero, stats counter, Why Choose Us, Specialists, Testimonials, Blog posts"),
        ("Frontend",      "src/frontend/src/pages/public/AboutPage.tsx",   "Chintu", "Completed", "Page hero, timeline, facility gallery, mission/vision/values"),
        ("Frontend",      "src/frontend/src/pages/public/DepartmentsPage.tsx","Chintu","Completed","Page hero, search filter, image cards with fallback"),
        ("Frontend",      "src/frontend/src/pages/public/DoctorProfilePage.tsx","Chintu","Completed","Two-column layout, booking CTA strip"),
        ("Frontend",      "src/frontend/src/pages/public/ContactPage.tsx", "Chintu", "Completed", "Split layout, icon info cards, map placeholder, success state"),
        ("Frontend",      "src/frontend/src/pages/public/LoginPage.tsx",   "Chintu", "Completed", "Split-screen, icon inputs, Eye/EyeOff password toggle"),
        ("Frontend",      "src/frontend/src/pages/public/SignupPage.tsx",  "Chintu", "Completed", "Split-screen, patient-only info banner, icon inputs"),
        ("Frontend",      "src/frontend/src/pages/admin/AdminDashboardPage.tsx","Chintu","Completed","Stat cards with icons + color borders + quick actions"),
        ("Frontend",      "src/frontend/public/images/ (28 SVGs)",         "Chintu", "Completed", "Placeholder images for heroes, facilities, departments, doctors"),
        # Config
        (".claude/agents/sunny.md",  "Agent Definition", "Akhil", "Completed", "Sunny scrum master agent definition"),
        ("docs/generate_tracker.py", "Script",           "Sunny", "Completed", "Python script that regenerates this Excel tracker"),
    ]

    for i, (cat, artifact, owner, status, desc) in enumerate(deliverables):
        bg = C_GREEN_LIGHT if i % 2 == 0 else C_WHITE
        _data_row(ws, row, [cat, artifact, owner, "", desc], bg=bg)
        _status_cell(ws, row, 4, status)
        ws.row_dimensions[row].height = 28
        row += 1

    _set_col_widths(ws, [16, 52, 12, 14, 60])


# ── Sheet 5: Bugs & Issues ───────────────────────────────────────────────────
def sheet_bugs(wb):
    ws = wb.create_sheet("Bugs & Issues")
    ws.sheet_view.showGridLines = False

    _title_block(ws, "🐛  Bugs & Issues Log",
                 "Logged by Gopal (QA). Updated after each test run.",
                 merge_cols=7)

    row = 4
    _header_row(ws, row,
                ["ID", "Date Found", "Severity", "Component", "Description", "Status", "Assigned To"],
                bg=C_GREEN_DARK)
    ws.row_dimensions[row].height = 24
    row += 1

    # Pre-populate known session-limit interruptions as informational items
    known_items = [
        ("INFO-001", datetime.now().strftime("%b %d, %Y"), "Info", "Pipeline",
         "Pavan and Chintu agents hit API session limit mid-work and were resumed. No data loss — both resumed cleanly from their transcript.",
         "Resolved", "Akhil"),
    ]

    for i, (bid, date, sev, comp, desc, status, assigned) in enumerate(known_items):
        bg = C_GREEN_LIGHT if i % 2 == 0 else C_WHITE
        _data_row(ws, row, [bid, date, sev, comp, desc, "", assigned], bg=bg)
        _status_cell(ws, row, 6, status)
        ws.row_dimensions[row].height = 36
        row += 1

    # Empty rows for Gopal to fill
    for i in range(15):
        bg = C_GREEN_LIGHT if (len(known_items) + i) % 2 == 0 else C_WHITE
        _data_row(ws, row, ["", "", "", "", "", "", ""], bg=bg)
        ws.row_dimensions[row].height = 28
        row += 1

    _set_col_widths(ws, [12, 14, 12, 18, 55, 14, 16])


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()

    sheet_overview(wb)
    sheet_agents(wb)
    sheet_scrum_log(wb)
    sheet_deliverables(wb)
    sheet_bugs(wb)

    out_path = os.path.join(os.path.dirname(__file__), "scrum-tracker.xlsx")
    wb.save(out_path)
    print(f"Scrum tracker saved -> {out_path}")


if __name__ == "__main__":
    main()
